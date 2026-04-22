from __future__ import annotations

import argparse
import heapq
import json
import math
import re
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any, Optional

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon as MplPolygon

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter  # noqa: F401
    XLSXWRITER_AVAILABLE = True
except Exception:
    XLSXWRITER_AVAILABLE = False


TRACTS_URL_BASE = "https://egisdata.baltimorecity.gov/egis/rest/services/311/ReferenceLayer/MapServer/49"

CITYLINK_COLOR_MAP = {
    "BLUE": "#0071BC",
    "BROWN": "#8B4513",
    "GOLD": "#DAA520",
    "GREEN": "#009E73",
    "LIME": "#9ACD32",
    "NAVY": "#1F3A93",
    "ORANGE": "#F58220",
    "PINK": "#D81B60",
    "PURPLE": "#6A3D9A",
    "RED": "#D62728",
    "SILVER": "#A7A9AC",
    "YELLOW": "#FFD200",
}


MTA_BUS_LINES_QUERY_URL = (
    "https://mdgeodata.md.gov/imap/rest/services/Transportation/MD_Transit/FeatureServer/10/query"
)

OSRM_ROUTE_BASE_URL = "https://router.project-osrm.org/route/v1/driving/"

CITYLINK_ABBREV_TO_NAME = {
    "BL": "CityLink BLUE",
    "BR": "CityLink BROWN",
    "GD": "CityLink GOLD",
    "GR": "CityLink GREEN",
    "LM": "CityLink LIME",
    "NV": "CityLink NAVY",
    "OR": "CityLink ORANGE",
    "PK": "CityLink PINK",
    "PR": "CityLink PURPLE",
    "RD": "CityLink RED",
    "SV": "CityLink SILVER",
    "YW": "CityLink YELLOW",
}

BALTIMORE_BBOX = {
    "min_lon": -76.93,
    "max_lon": -76.30,
    "min_lat": 39.08,
    "max_lat": 39.56,
}


def ensure_directory(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def safe_numeric(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.lower() == "nan":
        return None
    try:
        return float(s)
    except Exception:
        return None


def ask_path(prompt: str, default: Optional[Path] = None, allow_blank: bool = False) -> Optional[Path]:
    while True:
        suffix = f" [{default}]" if default else ""
        raw = input(f"{prompt}{suffix}: ").strip().strip('"')
        if raw == "" and allow_blank:
            return None
        candidate = Path(raw) if raw else default
        if candidate is None:
            print("Please enter a valid path.")
            continue
        if candidate.exists():
            return candidate
        print(f"Path not found: {candidate}")


def ask_output_dir(prompt: str, default: Optional[Path] = None) -> Path:
    suffix = f" [{default}]" if default else ""
    raw = input(f"{prompt}{suffix}: ").strip().strip('"')
    candidate = Path(raw) if raw else default
    if candidate is None:
        raise ValueError("No output directory supplied.")
    ensure_directory(candidate)
    return candidate


def read_any_table(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path, dtype=object)
    if suffix in {".xlsx", ".xls", ".xlsm"}:
        return pd.read_excel(path, sheet_name=sheet_name, dtype=object)
    raise ValueError(f"Unsupported file type: {path}")


def load_route_matrix(path: Path) -> tuple[pd.DataFrame, list[str], str]:
    """
    Supports route-by-tract formats like:
    - first column = candidate_route_id
    - first column = route_label
    - workbook with a sheet containing those columns
    """
    suffix = path.suffix.lower()
    chosen_sheet = ""

    if suffix in {".xlsx", ".xls", ".xlsm"}:
        xls = pd.ExcelFile(path)
        candidates = []
        for sheet in xls.sheet_names:
            df = pd.read_excel(path, sheet_name=sheet, dtype=object)
            df.columns = [str(c) for c in df.columns]
            first_col = str(df.columns[0]) if len(df.columns) else ""
            if first_col in {"candidate_route_id", "route_label"}:
                candidates.append((sheet, df))
        if not candidates:
            sheet = xls.sheet_names[0]
            df = pd.read_excel(path, sheet_name=sheet, dtype=object)
            chosen_sheet = sheet
        else:
            chosen_sheet, df = candidates[0]
    else:
        df = pd.read_csv(path, dtype=object)

    df.columns = [str(c) for c in df.columns]

    if "route_label" in df.columns:
        route_col = "route_label"
    elif "candidate_route_id" in df.columns:
        route_col = "candidate_route_id"
    else:
        route_col = df.columns[0]
        df = df.rename(columns={route_col: "route_label"})
        route_col = "route_label"

    if route_col != "route_label":
        df = df.rename(columns={route_col: "route_label"})

    # Drop empty / unnamed columns
    keep_cols = [c for c in df.columns if not str(c).startswith("Unnamed")]
    df = df[keep_cols].copy()

    tract_cols = []
    for c in df.columns:
        if c == "route_label":
            continue
        cs = str(c).strip()
        if re.fullmatch(r"\d{11}", cs):
            tract_cols.append(cs)

    if not tract_cols:
        tract_cols = [str(c) for c in df.columns if c != "route_label"]

    out = df[["route_label"] + tract_cols].copy()
    out["route_label"] = out["route_label"].astype(str).str.strip()

    # Prefer CityLink, but if everything is CityLink-like, keep all rows.
    citylink_mask = out["route_label"].str.contains("CITYLINK", case=False, na=False)
    if citylink_mask.any():
        out = out[citylink_mask].copy()

    return out.reset_index(drop=True), tract_cols, chosen_sheet


def load_distance_matrix(path: Path) -> pd.DataFrame:
    df = read_any_table(path)
    df.columns = [str(c) for c in df.columns]
    first_col = df.columns[0]
    df = df.rename(columns={first_col: "tract_id"})
    df["tract_id"] = df["tract_id"].astype(str).str.strip()
    df.columns = ["tract_id"] + [str(c).strip() for c in df.columns[1:]]
    df = df.set_index("tract_id")
    df.index = df.index.astype(str)
    return df


def sequence_from_row(row: pd.Series, tract_cols: list[str]) -> list[str]:
    seq_pairs = []
    for tract in tract_cols:
        v = safe_numeric(row.get(tract))
        if v is None or v <= 0:
            continue
        seq_pairs.append((int(v), str(tract)))
    seq_pairs.sort(key=lambda x: x[0])
    return [tract for _, tract in seq_pairs]


def build_graph(distance_df: pd.DataFrame) -> dict[str, dict[str, float]]:
    graph: dict[str, dict[str, float]] = {}
    for src in distance_df.index:
        graph.setdefault(str(src), {})
        for dst in distance_df.columns:
            if str(src) == str(dst):
                continue
            value = safe_numeric(distance_df.loc[src, dst])
            if value is None or value <= 0:
                continue
            graph[str(src)][str(dst)] = float(value)
    return graph


def shortest_path(graph: dict[str, dict[str, float]], start: str, goal: str) -> list[str]:
    if start == goal:
        return [start]

    pq = [(0.0, start)]
    dist = {start: 0.0}
    prev = {}
    visited = set()

    while pq:
        d, node = heapq.heappop(pq)
        if node in visited:
            continue
        visited.add(node)
        if node == goal:
            break
        for nbr, weight in graph.get(node, {}).items():
            nd = d + float(weight)
            if nd < dist.get(nbr, float("inf")):
                dist[nbr] = nd
                prev[nbr] = node
                heapq.heappush(pq, (nd, nbr))

    if goal not in dist:
        return []

    path = [goal]
    cur = goal
    while cur != start:
        cur = prev[cur]
        path.append(cur)
    path.reverse()
    return path


def route_color(route_label: str) -> str:
    upper = route_label.upper()
    for name, color in CITYLINK_COLOR_MAP.items():
        if name in upper:
            return color
    return "#555555"


def canonical_route_name(route_label: str) -> str:
    label = str(route_label).strip()
    return label.replace("—", "-").strip()


def build_route_bundle(route_df: pd.DataFrame, tract_cols: list[str], network_name: str) -> pd.DataFrame:
    rows = []
    for _, row in route_df.iterrows():
        label = str(row["route_label"]).strip()
        seq = sequence_from_row(row, tract_cols)
        if not seq:
            continue
        out = {
            "network_name": network_name,
            "route_label": label,
            "route_header": f"{network_name} | {canonical_route_name(label)}",
            "display_color": route_color(label),
            "origin_geoid": seq[0],
            "ending_geoid": seq[-1],
            "num_tracts": len(seq),
        }
        for tract in tract_cols:
            v = safe_numeric(row.get(tract))
            out[tract] = 0 if v is None else int(v)
        rows.append(out)
    return pd.DataFrame(rows)


def compute_optimized_from_distance(current_bundle_df: pd.DataFrame, tract_cols: list[str], distance_df: pd.DataFrame) -> pd.DataFrame:
    graph = build_graph(distance_df)
    rows = []

    for _, row in current_bundle_df.iterrows():
        current_seq = sequence_from_row(row, tract_cols)
        if len(current_seq) < 2:
            continue
        origin = current_seq[0]
        destination = current_seq[-1]
        optimized_seq = shortest_path(graph, origin, destination)
        if len(optimized_seq) < 2:
            optimized_seq = current_seq[:]

        out = {
            "network_name": "Optimized",
            "route_label": row["route_label"],
            "route_header": f"Optimized | {canonical_route_name(str(row['route_label']))}",
            "display_color": row["display_color"],
            "origin_geoid": origin,
            "ending_geoid": destination,
            "num_tracts": len(optimized_seq),
        }
        for tract in tract_cols:
            out[tract] = 0
        for i, tract in enumerate(optimized_seq, start=1):
            if tract in out:
                out[tract] = i
        rows.append(out)

    return pd.DataFrame(rows)


def matrix_for_excel(bundle_df: pd.DataFrame, tract_cols: list[str]) -> pd.DataFrame:
    if bundle_df.empty:
        return pd.DataFrame(columns=["route_label"] + tract_cols)
    out = bundle_df[["route_header"] + tract_cols].copy()
    out = out.rename(columns={"route_header": "route_label"})
    return out.reset_index(drop=True)


def tract_geojson_url() -> str:
    params = {
        "where": "1=1",
        "outFields": "*",
        "returnGeometry": "true",
        "f": "geojson",
        "outSR": "4326",
    }
    return f"{TRACTS_URL_BASE}/query?{urllib.parse.urlencode(params)}"


def download_json(url: str) -> dict[str, Any]:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


def extract_polygon_points(geometry: dict[str, Any]) -> list[list[tuple[float, float]]]:
    gtype = geometry.get("type")
    coords = geometry.get("coordinates", [])
    polys = []
    if gtype == "Polygon":
        if coords:
            polys.append([(float(x), float(y)) for x, y in coords[0]])
    elif gtype == "MultiPolygon":
        for poly in coords:
            if poly:
                polys.append([(float(x), float(y)) for x, y in poly[0]])
    return polys


def centroid(points: list[tuple[float, float]]) -> tuple[float, float]:
    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    return (sum(xs) / len(xs), sum(ys) / len(ys)) if points else (0.0, 0.0)


def load_tract_geometries() -> tuple[dict[str, tuple[float, float]], list[list[tuple[float, float]]]]:
    gj = download_json(tract_geojson_url())
    centroids = {}
    tract_polys = []
    for feat in gj.get("features", []):
        props = feat.get("properties", {}) or {}
        geoid = str(props.get("GEOID") or props.get("GEOID20") or props.get("GEOID10") or props.get("TRACTCE") or "").strip()
        polys = extract_polygon_points(feat.get("geometry", {}) or {})
        tract_polys.extend(polys)
        if geoid and polys:
            biggest = max(polys, key=len)
            centroids[geoid] = centroid(biggest)
    return centroids, tract_polys


def extract_line_parts(geometry: dict[str, Any]) -> list[list[tuple[float, float]]]:
    gtype = geometry.get("type")
    coords = geometry.get("coordinates", [])
    parts: list[list[tuple[float, float]]] = []
    if gtype == "LineString":
        if coords:
            parts.append([(float(x), float(y)) for x, y in coords])
    elif gtype == "MultiLineString":
        for part in coords:
            if part:
                parts.append([(float(x), float(y)) for x, y in part])
    return parts


def point_in_baltimore(x: float, y: float) -> bool:
    return (
        BALTIMORE_BBOX["min_lon"] <= x <= BALTIMORE_BBOX["max_lon"]
        and BALTIMORE_BBOX["min_lat"] <= y <= BALTIMORE_BBOX["max_lat"]
    )


def coords_near_baltimore(coords: list[tuple[float, float]]) -> bool:
    if not coords:
        return False
    return any(point_in_baltimore(x, y) for x, y in coords)


def geometry_near_baltimore(geometry: dict[str, Any]) -> bool:
    gtype = geometry.get("type")
    if gtype in ("LineString", "MultiLineString"):
        for part in extract_line_parts(geometry):
            if coords_near_baltimore(part):
                return True
        return False
    if gtype == "Point":
        coords = geometry.get("coordinates", [])
        return bool(coords) and point_in_baltimore(float(coords[0]), float(coords[1]))
    return False


def infer_citylink_key(text_value: str) -> str:
    upper = str(text_value or "").upper().strip()
    if not upper:
        return ""
    for abbr, full_name in CITYLINK_ABBREV_TO_NAME.items():
        if re.search(rf"(?<![A-Z]){re.escape(abbr)}(?![A-Z])", upper):
            return full_name
        color_name = full_name.replace("CityLink ", "").upper()
        if color_name in upper:
            return full_name
        if full_name.upper() in upper:
            return full_name
    return ""


def build_bundle_citylink_keys(bundle_df: pd.DataFrame) -> set[str]:
    keys = set()
    for _, row in bundle_df.iterrows():
        for candidate in [row.get("route_label", ""), row.get("route_header", "")]:
            key = infer_citylink_key(str(candidate))
            if key:
                keys.add(key)
    return keys


def fetch_live_mta_routes_geojson() -> dict[str, Any]:
    params = {
        "where": "1=1",
        "outFields": "OBJECTID,Route_Name,Route_Type,Route_Number,Distribution_Policy,Shape__Length",
        "returnGeometry": "true",
        "outSR": "4326",
        "f": "geojson",
    }
    url = MTA_BUS_LINES_QUERY_URL + "?" + urllib.parse.urlencode(params)
    return download_json(url)


def feature_citylink_key(feature: dict[str, Any]) -> str:
    props = feature.get("properties", {}) or {}
    for candidate in [props.get("Route_Number", ""), props.get("Route_Name", ""), props.get("Route_Type", "")]:
        key = infer_citylink_key(str(candidate))
        if key:
            return key
    return ""


def filter_live_mta_citylink_features(live_geojson: dict[str, Any], allowed_keys: set[str]) -> list[dict[str, Any]]:
    matches: list[dict[str, Any]] = []
    for feature in live_geojson.get("features", []):
        key = feature_citylink_key(feature)
        geometry = feature.get("geometry", {}) or {}
        if key and key in allowed_keys and geometry_near_baltimore(geometry):
            matches.append(feature)
    return matches


def clean_route_text(value: Any) -> str:
    value = str(value or "").strip()
    value = re.sub(r"\s+", " ", value)
    return value


def bundle_route_key(row: pd.Series) -> str:
    for candidate in [row.get("route_label", ""), row.get("route_header", "")]:
        key = infer_citylink_key(str(candidate))
        if key:
            return key
    return ""


def bundle_descriptor_text(row: pd.Series) -> str:
    raw = clean_route_text(row.get("route_label", "") or row.get("route_header", ""))
    key = bundle_route_key(row)
    raw = re.sub(r"^(Current|Optimized)\s*\|\s*", "", raw, flags=re.I).strip()
    if key:
        raw = re.sub(rf"^{re.escape(key)}\s*[-|]\s*", "", raw, flags=re.I).strip()
        raw = raw.replace(key, "").strip(" |-")
    return clean_route_text(raw)


def geometry_vertex_count(geometry: dict[str, Any]) -> int:
    gtype = geometry.get("type", "")
    coords = geometry.get("coordinates", [])
    if gtype == "LineString":
        return len(coords)
    if gtype == "MultiLineString":
        return sum(len(part) for part in coords if isinstance(part, list))
    if gtype == "Point":
        return 1 if coords else 0
    return 0


def dedupe_live_mta_citylink_features(
    matched_features: list[dict[str, Any]],
    current_bundle_df: pd.DataFrame,
) -> list[dict[str, Any]]:
    expected_by_key: dict[str, str] = {}
    order: list[str] = []

    for _, row in current_bundle_df.iterrows():
        key = bundle_route_key(row)
        if key and key not in expected_by_key:
            expected_by_key[key] = bundle_descriptor_text(row).upper()
            order.append(key)

    grouped: dict[str, list[dict[str, Any]]] = {}
    for feature in matched_features:
        key = feature_citylink_key(feature)
        if key:
            grouped.setdefault(key, []).append(feature)

    chosen: list[dict[str, Any]] = []

    for key in order:
        candidates = grouped.get(key, [])
        if not candidates:
            continue

        expected = expected_by_key.get(key, "").upper()

        def score(feature: dict[str, Any]) -> tuple[int, int]:
            props = feature.get("properties", {}) or {}
            route_name = clean_route_text(props.get("Route_Name", "")).upper()
            route_number = clean_route_text(props.get("Route_Number", "")).upper()
            score_value = 0

            if expected:
                if route_name == expected:
                    score_value += 300
                elif expected and (expected in route_name or route_name in expected):
                    score_value += 180
                else:
                    exp_words = set(re.findall(r"[A-Z0-9]+", expected))
                    name_words = set(re.findall(r"[A-Z0-9]+", route_name))
                    overlap = len((exp_words & name_words) - {"CITYLINK", "CURRENT", "OPTIMIZED"})
                    score_value += overlap * 20

            key_upper = key.upper()
            key_short = key_upper.replace("CITYLINK ", "")
            if route_name == key_upper or route_name == key_short:
                score_value += 35
            if route_number == key_upper or route_number == key_short:
                score_value += 45

            if re.fullmatch(r"\d+", route_number):
                score_value -= 25
            if re.fullmatch(r"\d+", route_name):
                score_value -= 20

            return score_value, geometry_vertex_count(feature.get("geometry", {}) or {})

        chosen.append(max(candidates, key=score))

    return chosen


def unique_bundle_by_route_key(bundle_df: pd.DataFrame) -> pd.DataFrame:
    seen: set[str] = set()
    keep_idx: list[int] = []
    for idx, row in bundle_df.iterrows():
        key = bundle_route_key(row)
        if not key or key in seen:
            continue
        seen.add(key)
        keep_idx.append(idx)
    return bundle_df.loc[keep_idx].copy()


def align_bundle_route_sets(
    current_bundle_df: pd.DataFrame,
    optimized_bundle_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    current_unique = unique_bundle_by_route_key(current_bundle_df)
    optimized_unique = unique_bundle_by_route_key(optimized_bundle_df)

    current_keys = [bundle_route_key(row) for _, row in current_unique.iterrows()]
    optimized_map = {bundle_route_key(row): row for _, row in optimized_unique.iterrows()}

    missing_optimized = [k for k in current_keys if k not in optimized_map]
    extra_optimized = [k for k in optimized_map.keys() if k not in set(current_keys)]

    if missing_optimized or extra_optimized:
        raise ValueError(
            "Current and optimized centroid route sets do not match. "
            f"Missing optimized routes: {missing_optimized}. "
            f"Extra optimized routes: {extra_optimized}."
        )

    ordered_opt_rows = [optimized_map[k] for k in current_keys]
    optimized_aligned = pd.DataFrame(ordered_opt_rows).reset_index(drop=True)
    current_aligned = current_unique.reset_index(drop=True)

    return current_aligned, optimized_aligned


def route_keys_from_geojson(feature_collection: dict[str, Any]) -> list[str]:
    keys: list[str] = []
    for feature in feature_collection.get("features", []):
        props = feature.get("properties", {}) or {}
        key = infer_citylink_key(str(props.get("route_label", "")) or str(props.get("route_header", "")))
        if key:
            keys.append(key)
    return keys


def validate_current_optimized_route_counts(
    current_bundle_df: pd.DataFrame,
    optimized_bundle_df: pd.DataFrame,
    current_live_street_geojson: dict[str, Any],
    optimized_simulated_geojson: dict[str, Any],
) -> tuple[pd.DataFrame, list[str]]:
    current_centroid_keys = [bundle_route_key(row) for _, row in unique_bundle_by_route_key(current_bundle_df).iterrows()]
    optimized_centroid_keys = [bundle_route_key(row) for _, row in unique_bundle_by_route_key(optimized_bundle_df).iterrows()]
    current_street_keys = route_keys_from_geojson(current_live_street_geojson)
    optimized_street_keys = route_keys_from_geojson(optimized_simulated_geojson)

    summary_df = pd.DataFrame(
        [
            {"layer_type": "centroid_to_centroid", "network": "current", "route_count": len(current_centroid_keys), "route_keys": " | ".join(current_centroid_keys)},
            {"layer_type": "centroid_to_centroid", "network": "optimized", "route_count": len(optimized_centroid_keys), "route_keys": " | ".join(optimized_centroid_keys)},
            {"layer_type": "street_to_street", "network": "current", "route_count": len(current_street_keys), "route_keys": " | ".join(current_street_keys)},
            {"layer_type": "street_to_street", "network": "optimized", "route_count": len(optimized_street_keys), "route_keys": " | ".join(optimized_street_keys)},
        ]
    )

    errors: list[str] = []
    if len(current_centroid_keys) != len(optimized_centroid_keys) or set(current_centroid_keys) != set(optimized_centroid_keys):
        errors.append(
            "Centroid current vs optimized route sets do not match. "
            f"Current={current_centroid_keys}; Optimized={optimized_centroid_keys}"
        )
    if len(current_street_keys) != len(optimized_street_keys) or set(current_street_keys) != set(optimized_street_keys):
        errors.append(
            "Street current vs optimized route sets do not match. "
            f"Current={current_street_keys}; Optimized={optimized_street_keys}"
        )

    return summary_df, errors


def bundle_route_to_citylink_key(bundle_row: pd.Series) -> str:
    for candidate in [bundle_row.get("route_label", ""), bundle_row.get("route_header", "")]:
        key = infer_citylink_key(str(candidate))
        if key:
            return key
    return ""


def densify_path(coords: list[tuple[float, float]], steps_per_segment: int = 8) -> list[tuple[float, float]]:
    if len(coords) < 2:
        return coords[:]
    out: list[tuple[float, float]] = []
    for i in range(len(coords) - 1):
        x0, y0 = coords[i]
        x1, y1 = coords[i + 1]
        for step in range(steps_per_segment):
            t = step / float(steps_per_segment)
            out.append((x0 + (x1 - x0) * t, y0 + (y1 - y0) * t))
    out.append(coords[-1])
    return out


def simulated_route_coords(bundle_row: pd.Series, tract_cols: list[str], centroid_lookup: dict[str, tuple[float, float]]) -> list[tuple[float, float]]:
    coords = line_coords_from_route(bundle_row, tract_cols, centroid_lookup)
    if len(coords) < 2:
        return coords
    coords = smooth_offset(coords, amplitude=0.0018, offset=0.0015)
    return densify_path(coords, steps_per_segment=10)


def _request_osrm_route(coords: list[tuple[float, float]]) -> list[tuple[float, float]]:
    if len(coords) < 2:
        return coords[:]

    cleaned: list[tuple[float, float]] = []
    for coord in coords:
        if not cleaned or coord != cleaned[-1]:
            cleaned.append(coord)
    if len(cleaned) < 2:
        return cleaned

    coord_str = ";".join(f"{x:.6f},{y:.6f}" for x, y in cleaned)
    params = urllib.parse.urlencode({
        "overview": "full",
        "geometries": "geojson",
        "steps": "false",
        "continue_straight": "false",
    })
    url = f"{OSRM_ROUTE_BASE_URL}{coord_str}?{params}"

    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        routes = data.get("routes") or []
        if not routes:
            return cleaned
        route_coords = (routes[0].get("geometry", {}) or {}).get("coordinates") or []
        snapped = [(float(x), float(y)) for x, y in route_coords]
        return snapped if len(snapped) >= 2 else cleaned
    except Exception:
        return cleaned


def road_route_coords_from_centroids(
    bundle_row: pd.Series,
    tract_cols: list[str],
    centroid_lookup: dict[str, tuple[float, float]],
    max_chunk_points: int = 20,
) -> list[tuple[float, float]]:
    centroid_coords = line_coords_from_route(bundle_row, tract_cols, centroid_lookup)
    if len(centroid_coords) < 2:
        return centroid_coords

    if len(centroid_coords) <= max_chunk_points:
        return _request_osrm_route(centroid_coords)

    stitched: list[tuple[float, float]] = []
    start = 0
    while start < len(centroid_coords) - 1:
        end = min(start + max_chunk_points, len(centroid_coords))
        chunk = centroid_coords[start:end]
        if len(chunk) < 2:
            break
        snapped_chunk = _request_osrm_route(chunk)
        if stitched and snapped_chunk and stitched[-1] == snapped_chunk[0]:
            snapped_chunk = snapped_chunk[1:]
        stitched.extend(snapped_chunk)
        if end >= len(centroid_coords):
            break
        start = end - 1

    return stitched if len(stitched) >= 2 else centroid_coords

def plot_live_mta_current_routes_png(
    current_bundle_df: pd.DataFrame,
    tract_polys: list[list[tuple[float, float]]],
    output_path: Path,
    title: str,
    exclude_silver_315: bool = False,
) -> None:
    fig, ax = plt.subplots(figsize=(18, 14))
    for poly in tract_polys:
        ax.add_patch(MplPolygon(poly, closed=True, facecolor="#E6E6E6", edgecolor="#9A9A9A", linewidth=0.35, alpha=0.8))

    live_geojson = fetch_live_mta_routes_geojson()
    allowed_keys = build_bundle_citylink_keys(current_bundle_df)
    matched_features = filter_live_mta_citylink_features(live_geojson, allowed_keys)

    seen_labels = set()
    legend_handles = []
    legend_labels = []

    for feature in matched_features:
        props = feature.get("properties", {}) or {}
        route_key = feature_citylink_key(feature)
        route_name_raw = str(props.get("Route_Name", "")).upper()
        route_number_raw = str(props.get("Route_Number", "")).upper()
        if exclude_silver_315 and (str(route_key).upper() == "CITYLINK SILVER") and ("COLUMBIA & SILVER SPRING - DC" in route_name_raw) and (route_number_raw == "315"):
            continue
        color = route_color(route_key or props.get("Route_Name", ""))
        parts = extract_line_parts(feature.get("geometry", {}) or {})
        for part in parts:
            if len(part) < 2:
                continue
            xs = [x for x, _ in part]
            ys = [y for _, y in part]
            line, = ax.plot(xs, ys, linewidth=2.4, color=color, alpha=0.95)
            label = route_key or str(props.get("Route_Name", "Unknown Route"))
            if label not in seen_labels:
                seen_labels.add(label)
                legend_handles.append(line)
                legend_labels.append(label)

    if legend_handles:
        ax.legend(
            legend_handles,
            legend_labels,
            loc="upper left",
            bbox_to_anchor=(1.02, 1.0),
            fontsize=8,
            frameon=True,
            title="Live MTA current routes",
            borderaxespad=0.0,
        )

    ax.set_title(title)
    ax.set_aspect("equal", adjustable="datalim")
    ax.axis("off")
    plt.tight_layout()

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(str(output_path.resolve()), dpi=300, bbox_inches="tight")

    plt.close(fig)


def plot_simulated_optimized_routes_png(
    optimized_bundle_df: pd.DataFrame,
    tract_cols: list[str],
    centroid_lookup: dict[str, tuple[float, float]],
    tract_polys: list[list[tuple[float, float]]],
    output_path: Path,
    title: str,
) -> None:
    fig, ax = plt.subplots(figsize=(18, 14))
    for poly in tract_polys:
        ax.add_patch(MplPolygon(poly, closed=True, facecolor="#E6E6E6", edgecolor="#9A9A9A", linewidth=0.35, alpha=0.8))

    legend_handles = []
    legend_labels = []

    for _, row in optimized_bundle_df.iterrows():
        coords = road_route_coords_from_centroids(row, tract_cols, centroid_lookup)
        if len(coords) < 2:
            continue
        xs = [x for x, _ in coords]
        ys = [y for _, y in coords]
        line, = ax.plot(xs, ys, linewidth=2.4, color=row["display_color"], alpha=0.92)
        sample_step = max(1, len(xs) // 25)
        ax.scatter(xs[::sample_step], ys[::sample_step], s=10, color=row["display_color"], alpha=0.45, zorder=5)
        legend_handles.append(line)
        legend_labels.append(str(row.get("route_header", row.get("route_label", "Unknown Route"))))

    if legend_handles:
        ax.legend(
            legend_handles,
            legend_labels,
            loc="upper left",
            bbox_to_anchor=(1.02, 1.0),
            fontsize=8,
            frameon=True,
            title="Road-following optimized routes",
            borderaxespad=0.0,
        )

    ax.set_title(title)
    ax.set_aspect("equal", adjustable="datalim")
    ax.axis("off")
    plt.tight_layout()

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(str(output_path.resolve()), dpi=300, bbox_inches="tight")

    plt.close(fig)


def build_live_mta_current_geojson(current_bundle_df: pd.DataFrame) -> dict[str, Any]:
    live_geojson = fetch_live_mta_routes_geojson()
    allowed_keys = build_bundle_citylink_keys(current_bundle_df)
    matched_features = filter_live_mta_citylink_features(live_geojson, allowed_keys)
    matched_features = dedupe_live_mta_citylink_features(matched_features, current_bundle_df)
    features: list[dict[str, Any]] = []

    def _equivalent_to_route_key(value: str, route_key: str) -> bool:
        v = clean_route_text(value).upper()
        rk = clean_route_text(route_key).upper()
        if not v:
            return False
        if v == rk:
            return True
        return v == rk.replace("CITYLINK ", "")

    for feature in matched_features:
        props = feature.get("properties", {}) or {}
        route_key = feature_citylink_key(feature) or str(props.get("Route_Name", "Unknown Route"))
        color = route_color(route_key)
        route_name_raw = clean_route_text(props.get("Route_Name", ""))
        route_number_raw = clean_route_text(props.get("Route_Number", ""))

        route_header = f"Current Street | {route_key}"

        extra_parts = []
        if route_name_raw and not _equivalent_to_route_key(route_name_raw, route_key):
            extra_parts.append(route_name_raw)
        if route_number_raw and not _equivalent_to_route_key(route_number_raw, route_key):
            extra_parts.append(route_number_raw)

        if extra_parts:
            route_header = f"{route_header} | " + " | ".join(extra_parts)

        features.append(
            {
                "type": "Feature",
                "properties": {
                    "route_header": route_header,
                    "route_label": route_key,
                    "display_color": color,
                    "network_name": "Current Street",
                    "route_name_raw": route_name_raw,
                    "route_number_raw": route_number_raw,
                    "route_type_raw": str(props.get("Route_Type", "")),
                },
                "geometry": feature.get("geometry", {}),
            }
        )

    return {"type": "FeatureCollection", "features": features}


def build_simulated_optimized_geojson(
    optimized_bundle_df: pd.DataFrame,
    tract_cols: list[str],
    centroid_lookup: dict[str, tuple[float, float]],
) -> dict[str, Any]:
    feats: list[dict[str, Any]] = []
    for _, row in optimized_bundle_df.iterrows():
        coords = road_route_coords_from_centroids(row, tract_cols, centroid_lookup)
        if len(coords) < 2:
            continue
        if not coords_near_baltimore(coords):
            continue
        feats.append(
            {
                "type": "Feature",
                "properties": {
                    "route_header": f"Optimized Street | {row['route_label']}",
                    "route_label": row["route_label"],
                    "display_color": row["display_color"],
                    "network_name": "Optimized Street",
                    "origin_geoid": row["origin_geoid"],
                    "ending_geoid": row["ending_geoid"],
                    "num_tracts": row["num_tracts"],
                },
                "geometry": {
                    "type": "LineString",
                    "coordinates": [[x, y] for x, y in coords],
                },
            }
        )
    return {"type": "FeatureCollection", "features": feats}


def line_coords_from_route(bundle_row: pd.Series, tract_cols: list[str], centroid_lookup: dict[str, tuple[float, float]]) -> list[tuple[float, float]]:
    seq = sequence_from_row(bundle_row, tract_cols)
    return [centroid_lookup[t] for t in seq if t in centroid_lookup]


def audit_plottable_routes(
    bundle_df: pd.DataFrame,
    tract_cols: list[str],
    centroid_lookup: dict[str, tuple[float, float]],
) -> tuple[pd.DataFrame, list[str]]:
    audit_rows = []
    missing_routes: list[str] = []

    for _, row in bundle_df.iterrows():
        route_name = str(row.get("route_header", row.get("route_label", "Unknown Route")))
        seq = sequence_from_row(row, tract_cols)
        coords = [centroid_lookup[t] for t in seq if t in centroid_lookup]
        drawable = len(coords) >= 2

        audit_rows.append(
            {
                "route_name": route_name,
                "num_sequence_tracts": len(seq),
                "num_centroid_points_found": len(coords),
                "is_drawable_on_png": drawable,
            }
        )

        if not drawable:
            missing_routes.append(route_name)

    return pd.DataFrame(audit_rows), missing_routes


def filter_bundle_to_baltimore(
    bundle_df: pd.DataFrame,
    tract_cols: list[str],
    centroid_lookup: dict[str, tuple[float, float]],
) -> pd.DataFrame:
    if bundle_df.empty:
        return bundle_df.copy()

    keep_rows = []
    for _, row in bundle_df.iterrows():
        coords = line_coords_from_route(row, tract_cols, centroid_lookup)
        if len(coords) < 2:
            continue
        if coords_near_baltimore(coords):
            keep_rows.append(row.to_dict())

    return pd.DataFrame(keep_rows)


def smooth_offset(coords: list[tuple[float, float]], amplitude: float = 0.0018, offset: float = 0.0) -> list[tuple[float, float]]:
    if len(coords) < 2:
        return coords[:]
    # gentle wiggle through centroid centers
    out = []
    for i, (x, y) in enumerate(coords):
        if i == 0 or i == len(coords) - 1:
            out.append((x, y))
            continue
        x0, y0 = coords[i - 1]
        x1, y1 = coords[i + 1]
        dx, dy = x1 - x0, y1 - y0
        length = math.hypot(dx, dy)
        if length == 0:
            out.append((x, y))
            continue
        nx, ny = -dy / length, dx / length
        sign = 1 if i % 2 == 0 else -1
        out.append((x + sign * amplitude * nx, y + sign * amplitude * ny))
    if offset == 0:
        return out

    shifted = []
    for i, (x, y) in enumerate(out):
        if i == 0:
            x1, y1 = out[i + 1]
            dx, dy = x1 - x, y1 - y
        else:
            x0, y0 = out[i - 1]
            dx, dy = x - x0, y - y0
        length = math.hypot(dx, dy)
        if length == 0:
            shifted.append((x, y))
        else:
            nx, ny = -dy / length, dx / length
            shifted.append((x + offset * nx, y + offset * ny))
    return shifted


def plot_bundle_png(
    bundle_df: pd.DataFrame,
    tract_cols: list[str],
    centroid_lookup: dict[str, tuple[float, float]],
    tract_polys: list[list[tuple[float, float]]],
    output_path: Path,
    title: str,
    optimized_style: bool = False,
    route_alpha: float = 0.95,
    show_centroid_dots: bool = False,
    show_route_legend: bool = False,
) -> None:
    fig, ax = plt.subplots(figsize=(18, 14) if show_route_legend else (14, 14))

    for poly in tract_polys:
        ax.add_patch(MplPolygon(poly, closed=True, facecolor="#E6E6E6", edgecolor="#9A9A9A", linewidth=0.35, alpha=0.8))

    legend_handles = []
    legend_labels = []

    for _, row in bundle_df.iterrows():
        coords = line_coords_from_route(row, tract_cols, centroid_lookup)
        if len(coords) < 2:
            continue
        if optimized_style:
            coords = smooth_offset(coords, amplitude=0.0018, offset=0.0022)

        xs = [x for x, _ in coords]
        ys = [y for _, y in coords]
        line, = ax.plot(
            xs, ys,
            linewidth=2.3,
            color=row["display_color"],
            linestyle="--" if optimized_style else "-",
            alpha=route_alpha,
        )

        if show_centroid_dots:
            ax.scatter(xs, ys, s=16, color=row["display_color"], alpha=min(1.0, route_alpha + 0.15), zorder=5)

        if show_route_legend:
            legend_handles.append(line)
            legend_labels.append(str(row.get("route_header", row.get("route_label", "Unknown Route"))))

    if show_route_legend and legend_handles:
        dot_handle = None
        if show_centroid_dots:
            dot_handle = plt.Line2D([0], [0], marker='o', linestyle='None', markersize=6, color='#333333', label='Tract centroid')
            legend_handles = [dot_handle] + legend_handles
            legend_labels = ['Tract centroid'] + legend_labels

        ax.legend(
            legend_handles,
            legend_labels,
            loc="upper left",
            bbox_to_anchor=(1.02, 1.0),
            fontsize=8,
            frameon=True,
            title="Legend",
            borderaxespad=0.0,
        )

    ax.set_title(title)
    ax.set_aspect("equal", adjustable="datalim")
    ax.axis("off")
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)

def plot_combo_png(
    current_df: pd.DataFrame,
    optimized_df: pd.DataFrame,
    tract_cols: list[str],
    centroid_lookup: dict[str, tuple[float, float]],
    tract_polys: list[list[tuple[float, float]]],
    output_path: Path,
    route_alpha: float = 0.92,
    show_centroid_dots: bool = False,
    show_route_legend: bool = False,
) -> None:
    fig, ax = plt.subplots(figsize=(18, 14) if show_route_legend else (14, 14))

    for poly in tract_polys:
        ax.add_patch(
            MplPolygon(
                poly,
                closed=True,
                facecolor="#E6E6E6",
                edgecolor="#9A9A9A",
                linewidth=0.35,
                alpha=0.8,
            )
        )

    for _, row in current_df.iterrows():
        coords = line_coords_from_route(row, tract_cols, centroid_lookup)
        if len(coords) < 2:
            continue
        xs = [x for x, _ in coords]
        ys = [y for _, y in coords]
        ax.plot(xs, ys, linewidth=2.1, color=row["display_color"], linestyle="-", alpha=route_alpha)
        if show_centroid_dots:
            ax.scatter(xs, ys, s=16, color=row["display_color"], alpha=min(1.0, route_alpha + 0.15), zorder=5)

    for _, row in optimized_df.iterrows():
        coords = line_coords_from_route(row, tract_cols, centroid_lookup)
        if len(coords) < 2:
            continue
        coords = smooth_offset(coords, amplitude=0.0018, offset=0.0022)
        xs = [x for x, _ in coords]
        ys = [y for _, y in coords]
        ax.plot(xs, ys, linewidth=2.1, color=row["display_color"], linestyle="--", alpha=route_alpha)
        if show_centroid_dots:
            ax.scatter(xs, ys, s=16, color=row["display_color"], alpha=min(1.0, route_alpha + 0.15), zorder=5)

    if show_route_legend:
        base_handles = [
            plt.Line2D([0], [0], color="black", linestyle="-", linewidth=2.1, label="Current routes"),
            plt.Line2D([0], [0], color="black", linestyle="--", linewidth=2.1, label="Optimized routes"),
        ]
        base_labels = ["Current routes", "Optimized routes"]

        if show_centroid_dots:
            base_handles.append(
                plt.Line2D([0], [0], marker="o", linestyle="None", markersize=6, color="black", label="Tract centroid")
            )
            base_labels.append("Tract centroid")

        ax.legend(
            base_handles,
            base_labels,
            loc="upper left",
            bbox_to_anchor=(1.02, 1.0),
            fontsize=8,
            frameon=True,
            title="Legend",
            borderaxespad=0.0,
        )

    ax.set_title("Baltimore tract centroid-to-centroid routes")
    ax.set_aspect("equal", adjustable="datalim")
    ax.axis("off")
    plt.tight_layout()

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(str(output_path.resolve()), dpi=300, bbox_inches="tight")

    plt.close(fig)

def plot_street_combo_png(
    current_bundle_df: pd.DataFrame,
    optimized_bundle_df: pd.DataFrame,
    tract_cols: list[str],
    centroid_lookup: dict[str, tuple[float, float]],
    tract_polys: list[list[tuple[float, float]]],
    output_path: Path,
    title: str,
    route_alpha: float = 0.92,
    show_route_legend: bool = True,
    exclude_silver_315: bool = False,
) -> None:
    fig, ax = plt.subplots(figsize=(18, 14) if show_route_legend else (14, 14))

    for poly in tract_polys:
        ax.add_patch(
            MplPolygon(
                poly,
                closed=True,
                facecolor="#E6E6E6",
                edgecolor="#9A9A9A",
                linewidth=0.35,
                alpha=0.8,
            )
        )

    live_geojson = fetch_live_mta_routes_geojson()
    allowed_keys = build_bundle_citylink_keys(current_bundle_df)
    matched_features = filter_live_mta_citylink_features(live_geojson, allowed_keys)
    matched_features = dedupe_live_mta_citylink_features(matched_features, current_bundle_df)

    for feature in matched_features:
        props = feature.get("properties", {}) or {}
        route_key = feature_citylink_key(feature)
        route_name_raw = str(props.get("Route_Name", "")).upper()
        route_number_raw = str(props.get("Route_Number", "")).upper()

        if (
            exclude_silver_315
            and str(route_key).upper() == "CITYLINK SILVER"
            and "COLUMBIA & SILVER SPRING - DC" in route_name_raw
            and route_number_raw == "315"
        ):
            continue

        color = route_color(route_key or props.get("Route_Name", ""))
        parts = extract_line_parts(feature.get("geometry", {}) or {})

        for part in parts:
            if len(part) < 2:
                continue
            xs = [x for x, _ in part]
            ys = [y for _, y in part]
            ax.plot(
                xs,
                ys,
                linewidth=2.3,
                color=color,
                linestyle="-",
                alpha=route_alpha,
            )

    for _, row in optimized_bundle_df.iterrows():
        coords = road_route_coords_from_centroids(row, tract_cols, centroid_lookup)
        if len(coords) < 2:
            continue

        xs = [x for x, _ in coords]
        ys = [y for _, y in coords]

        ax.plot(
            xs,
            ys,
            linewidth=2.3,
            color=row["display_color"],
            linestyle="--",
            alpha=route_alpha,
        )

    if show_route_legend:
        legend_handles = [
            plt.Line2D([0], [0], color="black", linestyle="-", linewidth=2.3, label="Current routes"),
            plt.Line2D([0], [0], color="black", linestyle="--", linewidth=2.3, label="Optimized routes"),
        ]
        legend_labels = ["Current routes", "Optimized routes"]

        ax.legend(
            legend_handles,
            legend_labels,
            loc="upper left",
            bbox_to_anchor=(1.02, 1.0),
            fontsize=8,
            frameon=True,
            title="Legend",
            borderaxespad=0.0,
        )

    ax.set_title(title)
    ax.set_aspect("equal", adjustable="datalim")
    ax.axis("off")
    plt.tight_layout()

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(str(output_path.resolve()), dpi=300, bbox_inches="tight")

    plt.close(fig)

    # -------------------------
    # Current street-to-street
    # -------------------------
    live_geojson = fetch_live_mta_routes_geojson()
    allowed_keys = build_bundle_citylink_keys(current_bundle_df)
    matched_features = filter_live_mta_citylink_features(live_geojson, allowed_keys)
    matched_features = dedupe_live_mta_citylink_features(matched_features, current_bundle_df)

    for feature in matched_features:
        props = feature.get("properties", {}) or {}
        route_key = feature_citylink_key(feature)
        route_name_raw = str(props.get("Route_Name", "")).upper()
        route_number_raw = str(props.get("Route_Number", "")).upper()

        if (
            exclude_silver_315
            and str(route_key).upper() == "CITYLINK SILVER"
            and "COLUMBIA & SILVER SPRING - DC" in route_name_raw
            and route_number_raw == "315"
        ):
            continue

        color = route_color(route_key or props.get("Route_Name", ""))
        parts = extract_line_parts(feature.get("geometry", {}) or {})

        for part in parts:
            if len(part) < 2:
                continue
            xs = [x for x, _ in part]
            ys = [y for _, y in part]
            ax.plot(
                xs,
                ys,
                linewidth=2.3,
                color=color,
                linestyle="-",
                alpha=route_alpha,
            )

    # -------------------------
    # Optimized street-to-street
    # -------------------------
    for _, row in optimized_bundle_df.iterrows():
        coords = road_route_coords_from_centroids(row, tract_cols, centroid_lookup)
        if len(coords) < 2:
            continue

        xs = [x for x, _ in coords]
        ys = [y for _, y in coords]

        ax.plot(
            xs,
            ys,
            linewidth=2.3,
            color=row["display_color"],
            linestyle="--",
            alpha=route_alpha,
        )

    if show_route_legend:
        legend_handles = [
            plt.Line2D([0], [0], color="black", linestyle="-", linewidth=2.3, label="Current routes"),
            plt.Line2D([0], [0], color="black", linestyle="--", linewidth=2.3, label="Optimized routes"),
        ]
        legend_labels = ["Current routes", "Optimized routes"]

        ax.legend(
            legend_handles,
            legend_labels,
            loc="upper left",
            bbox_to_anchor=(1.02, 1.0),
            fontsize=8,
            frameon=True,
            title="Legend",
            borderaxespad=0.0,
        )

    ax.set_title(title)
    ax.set_aspect("equal", adjustable="datalim")
    ax.axis("off")
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)

def features_from_bundle(bundle_df: pd.DataFrame, tract_cols: list[str], centroid_lookup: dict[str, tuple[float, float]], optimized_style: bool = False) -> dict[str, Any]:
    feats = []
    for _, row in bundle_df.iterrows():
        coords = line_coords_from_route(row, tract_cols, centroid_lookup)
        if len(coords) < 2:
            continue
        if optimized_style:
            coords = smooth_offset(coords, amplitude=0.0018, offset=0.0022)
        feats.append(
            {
                "type": "Feature",
                "properties": {
                    "route_header": row["route_header"],
                    "route_label": row["route_label"],
                    "display_color": row["display_color"],
                    "network_name": row["network_name"],
                    "origin_geoid": row["origin_geoid"],
                    "ending_geoid": row["ending_geoid"],
                    "num_tracts": row["num_tracts"],
                },
                "geometry": {
                    "type": "LineString",
                    "coordinates": [[x, y] for x, y in coords],
                },
            }
        )
    return {"type": "FeatureCollection", "features": feats}


def build_centroid_centers_geojson(
    current_bundle_df: pd.DataFrame,
    optimized_bundle_df: pd.DataFrame,
    tract_cols: list[str],
    centroid_lookup: dict[str, tuple[float, float]],
) -> dict[str, Any]:
    used_geoids: set[str] = set()

    for bundle_df in [current_bundle_df, optimized_bundle_df]:
        for _, row in bundle_df.iterrows():
            seq = sequence_from_row(row, tract_cols)
            for geoid in seq:
                if geoid in centroid_lookup:
                    used_geoids.add(geoid)

    feats = []
    for geoid in sorted(used_geoids):
        x, y = centroid_lookup[geoid]
        if not point_in_baltimore(x, y):
            continue
        feats.append(
            {
                "type": "Feature",
                "properties": {
                    "tract_id": geoid,
                    "layer_name": "Centroid centers",
                },
                "geometry": {
                    "type": "Point",
                    "coordinates": [x, y],
                },
            }
        )

    return {"type": "FeatureCollection", "features": feats}



HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>Baltimore CityLink centroid routes</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" crossorigin="" />
<style>
html, body, #map {height:100%; width:100%; margin:0; padding:0; background:#ffffff; font-family:Arial,Helvetica,sans-serif;}
.box {position:absolute; z-index:1000; background:rgba(255,255,255,.96); border:1px solid #bbb; border-radius:8px; box-shadow:0 1px 6px rgba(0,0,0,.2);}
.drag {cursor:move; padding:8px 10px; font-weight:700; border-bottom:1px solid #ddd; background:#f5f5f5; border-radius:8px 8px 0 0; display:flex; align-items:center; justify-content:space-between; gap:10px;}
.drag-title {flex:1; min-width:0;}
.panel-toggle {cursor:pointer; border:1px solid #aaa; background:#fff; border-radius:4px; padding:1px 8px; font-size:12px;}
.body {padding:10px; font-size:12px; line-height:1.35;}
.box.minimized .body {display:none;}
#legend {top:12px; right:12px; width:330px; max-height:32vh; overflow:auto;}
#picker {top:calc(12px + 32vh + 14px); right:12px; width:330px; max-height:28vh; overflow:auto;}
#groupbox {top:calc(12px + 60vh + 28px); right:12px; width:330px;}
#title {top:12px; left:12px; max-width:430px;}
.swatch {display:inline-block; width:12px; height:12px; border:1px solid #666; margin-right:6px; vertical-align:middle;}
.route-line {display:inline-block; width:18px; border-top:3px solid #000; margin-right:6px; vertical-align:middle;}
.picker-row {display:flex; justify-content:space-between; align-items:center; gap:8px; margin:6px 0;}
.picker-row label {flex:1; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;}
.group-row {display:flex; align-items:center; gap:8px; margin:8px 0;}
.group-row input {transform:scale(1.1);}
.note {font-size:11px; color:#555;}
</style>
</head>
<body>
<div id="map"></div>

<div id="title" class="box">
  <div class="drag"><span class="drag-title">Baltimore CityLink centroid-to-centroid routes</span><button class="panel-toggle" data-target="title">−</button></div>
  <div class="body">
    Gray tract polygons with centroid-to-centroid CityLink routes.
    Current routes are solid lines. Optimized routes are dashed lines passing through tract centroid centers.
    Use the layer control for individual routes, the all-routes toggles for grouped visibility, and the color picker to change line colors.
  </div>
</div>

<div id="legend" class="box">
  <div class="drag"><span class="drag-title">Legend</span><button class="panel-toggle" data-target="legend">−</button></div>
  <div class="body" id="legendBody">Loading...</div>
</div>

<div id="picker" class="box">
  <div class="drag"><span class="drag-title">Color picker</span><button class="panel-toggle" data-target="picker">−</button></div>
  <div class="body" id="pickerBody">Loading...</div>
</div>

<div id="groupbox" class="box">
  <div class="drag"><span class="drag-title">All-routes toggles</span><button class="panel-toggle" data-target="groupbox">−</button></div>
  <div class="body">
    <div class="group-row">
      <input type="checkbox" id="toggleCurrent" checked />
      <label for="toggleCurrent">Show all current centroid routes</label>
    </div>
    <div class="group-row">
      <input type="checkbox" id="toggleOptimized" checked />
      <label for="toggleOptimized">Show all optimized centroid routes</label>
    </div>
    <div class="note">These toggles work on top of the individual Leaflet layer controls.</div>
  </div>
</div>

<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js" crossorigin=""></script>
<script>
const TRACTS_URL = "__TRACTS_URL__";
const CURRENT_GEOJSON = __CURRENT_GEOJSON__;
const OPTIMIZED_GEOJSON = __OPTIMIZED_GEOJSON__;

const map = L.map('map').setView([39.29,-76.61], 11);
const baseLayer = L.tileLayer('https://{s}.basemaps.cartocdn.com/light_nolabels/{z}/{x}/{y}{r}.png', {
  attribution: '&copy; OpenStreetMap &copy; CARTO', subdomains:'abcd', maxZoom:20
}).addTo(map);

const routeRegistry = new Map();
const legendBody = document.getElementById('legendBody');
const pickerBody = document.getElementById('pickerBody');

function makeDraggable(el){
  const handle = el.querySelector('.drag') || el;
  let dragging = false, sx=0, sy=0, sl=0, st=0;
  handle.addEventListener('mousedown', (e)=>{
    if (e.target.classList.contains('panel-toggle')) return;
    const r = el.getBoundingClientRect();
    dragging = true; sx = e.clientX; sy = e.clientY; sl = r.left; st = r.top;
    el.style.left = `${r.left}px`; el.style.top = `${r.top}px`; el.style.right='auto'; el.style.bottom='auto';
    e.preventDefault();
  });
  window.addEventListener('mousemove', (e)=>{
    if(!dragging) return;
    el.style.left = `${Math.max(0, sl + e.clientX - sx)}px`;
    el.style.top = `${Math.max(0, st + e.clientY - sy)}px`;
  });
  window.addEventListener('mouseup', ()=> dragging = false);
}

function bindPanelMinimizeButtons() {
  document.querySelectorAll('.panel-toggle').forEach(btn => {
    btn.addEventListener('click', () => {
      const targetId = btn.getAttribute('data-target');
      const panel = document.getElementById(targetId);
      if (!panel) return;
      panel.classList.toggle('minimized');
      btn.textContent = panel.classList.contains('minimized') ? '+' : '−';
    });
  });
}

async function fetchJson(url){
  const r = await fetch(url);
  if(!r.ok) throw new Error(`Failed to fetch ${url}`);
  return await r.json();
}

function addLayerGroup(collection, groupName, prefix, dashArray){
  const overlay = {};
  (collection.features || []).forEach(feature=>{
    const props = feature.properties || {};
    const name = props.route_header || `${prefix} | Unknown`;
    const color = props.display_color || '#555555';
    const layer = L.geoJSON(feature, {
      style: {color: color, weight: 4, opacity: 0.96, dashArray: dashArray || null},
      onEachFeature: (feat, lyr) => {
        const p = feat.properties || {};
        lyr.bindPopup(`<b>${name}</b><br>Origin tract: ${p.origin_geoid || ''}<br>Destination tract: ${p.ending_geoid || ''}<br>Tracts in route: ${p.num_tracts || ''}`);
      }
    });
    overlay[name] = layer;
    routeRegistry.set(name, {layer, color, dashArray, group: groupName});
  });
  return overlay;
}

function setGroupVisibility(groupName, visible){
  routeRegistry.forEach((meta) => {
    if (meta.group !== groupName) return;
    if (visible) {
      if (!map.hasLayer(meta.layer)) meta.layer.addTo(map);
    } else {
      if (map.hasLayer(meta.layer)) map.removeLayer(meta.layer);
    }
  });
}

function buildLegend() {
  let html = `<div><span class="swatch" style="background:#E6E6E6"></span>Tract polygons</div>`;
  html += `<div><span class="route-line" style="border-top-color:#000"></span>Current route</div>`;
  html += `<div><span class="route-line" style="border-top-color:#000;border-top-style:dashed"></span>Optimized route</div><br>`;
  [...routeRegistry.entries()].sort((a,b)=>a[0].localeCompare(b[0])).forEach(([name, meta])=>{
    const style = meta.dashArray ? 'border-top-style:dashed;' : '';
    html += `<div><span class="route-line" style="border-top-color:${meta.color};${style}"></span>${name}</div>`;
  });
  legendBody.innerHTML = html;
}

function buildPicker(){
  let html = '';
  [...routeRegistry.entries()].sort((a,b)=>a[0].localeCompare(b[0])).forEach(([name, meta], idx)=>{
    html += `<div class="picker-row"><label for="p_${idx}" title="${name}">${name}</label><input type="color" id="p_${idx}" value="${meta.color}"></div>`;
  });
  pickerBody.innerHTML = html;
  [...routeRegistry.entries()].sort((a,b)=>a[0].localeCompare(b[0])).forEach(([name, meta], idx)=>{
    document.getElementById(`p_${idx}`).addEventListener('input', (e)=>{
      meta.color = e.target.value;
      meta.layer.eachLayer(l=>{
        if(l.setStyle){ l.setStyle({color: meta.color, dashArray: meta.dashArray || null}); }
      });
      buildLegend();
    });
  });
}

async function init(){
  const tracts = await fetchJson(TRACTS_URL);
  const tractLayer = L.geoJSON(tracts, {
    style: {color:'#808080', weight:1, opacity:0.85, fillColor:'#E6E6E6', fillOpacity:0.55},
    onEachFeature: (feature, layer) => {
      const props = feature.properties || {};
      const tractId = props.GEOID || props.GEOID20 || props.GEOID10 || props.TRACTCE || 'Unknown tract';
      layer.bindTooltip(`Tract ID: ${tractId}`, {sticky:true});
    }
  }).addTo(map);

  const overlays = {'Tracts': tractLayer};
  Object.assign(overlays, addLayerGroup(CURRENT_GEOJSON, 'current', 'Current', null));
  Object.assign(overlays, addLayerGroup(OPTIMIZED_GEOJSON, 'optimized', 'Optimized', '8 6'));

  L.control.layers({'Light': baseLayer}, overlays, {collapsed:true}).addTo(map);

  const bounds = tractLayer.getBounds();
  if(bounds.isValid()) map.fitBounds(bounds.pad(0.03));

  buildLegend();
  buildPicker();

  setGroupVisibility('current', document.getElementById('toggleCurrent').checked);
  setGroupVisibility('optimized', document.getElementById('toggleOptimized').checked);

  document.getElementById('toggleCurrent').addEventListener('change', (e) => {
    setGroupVisibility('current', e.target.checked);
  });
  document.getElementById('toggleOptimized').addEventListener('change', (e) => {
    setGroupVisibility('optimized', e.target.checked);
  });
}

makeDraggable(document.getElementById('legend'));
makeDraggable(document.getElementById('picker'));
makeDraggable(document.getElementById('groupbox'));
makeDraggable(document.getElementById('title'));
bindPanelMinimizeButtons();
init().catch(err => console.error(err));
</script>
</body>
</html>
"""

STREET_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>Baltimore CityLink street and centroid routes</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" crossorigin="" />
<style>
html, body, #map {height:100%; width:100%; margin:0; padding:0; background:#ffffff; font-family:Arial,Helvetica,sans-serif;}
.box {position:absolute; z-index:1000; background:rgba(255,255,255,.96); border:1px solid #bbb; border-radius:8px; box-shadow:0 1px 6px rgba(0,0,0,.2);}
.drag {cursor:move; padding:8px 10px; font-weight:700; border-bottom:1px solid #ddd; background:#f5f5f5; border-radius:8px 8px 0 0; display:flex; align-items:center; justify-content:space-between; gap:10px;}
.drag-title {flex:1; min-width:0;}
.panel-toggle {cursor:pointer; border:1px solid #aaa; background:#fff; border-radius:4px; padding:1px 8px; font-size:12px;}
.body {padding:10px; font-size:12px; line-height:1.35;}
.box.minimized .body {display:none;}
#legend {top:12px; right:12px; width:360px; max-height:26vh; overflow:auto;}
#picker {top:calc(12px + 26vh + 14px); right:12px; width:360px; max-height:24vh; overflow:auto;}
#groupbox {top:calc(12px + 50vh + 28px); right:12px; width:360px; max-height:38vh; overflow:auto;}
#title {top:12px; left:12px; max-width:500px;}
.swatch {display:inline-block; width:12px; height:12px; border:1px solid #666; margin-right:6px; vertical-align:middle;}
.route-line {display:inline-block; width:18px; border-top:3px solid #000; margin-right:6px; vertical-align:middle;}
.centroid-dot {display:inline-block; width:10px; height:10px; border-radius:50%; background:#111; border:1px solid #111; margin-right:8px; vertical-align:middle;}
.picker-row {display:flex; justify-content:space-between; align-items:center; gap:8px; margin:6px 0;}
.picker-row label {flex:1; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;}
.group-row {display:flex; align-items:center; gap:8px; margin:8px 0;}
.group-row input {transform:scale(1.1);}
.note {font-size:11px; color:#555;}
hr {border:none; border-top:1px solid #ddd; margin:8px 0;}
</style>
</head>
<body>
<div id="map"></div>

<div id="title" class="box">
  <div class="drag"><span class="drag-title">Baltimore CityLink street and centroid routes</span><button class="panel-toggle" data-target="title">−</button></div>
  <div class="body">
    This street-view HTML includes:
    current centroid-to-centroid routes, optimized centroid-to-centroid routes,
    current street-to-street routes, optimized street-to-street simulated routes,
    and a toggle for centroid centers.
    It also includes 9 basemap choices, individual layer toggles, and group toggles. If you turn off all route toggles, the map can be blank aside from the chosen basemap and any layers you keep on.
  </div>
</div>

<div id="legend" class="box">
  <div class="drag"><span class="drag-title">Legend</span><button class="panel-toggle" data-target="legend">−</button></div>
  <div class="body" id="legendBody">Loading...</div>
</div>

<div id="picker" class="box">
  <div class="drag"><span class="drag-title">Color picker</span><button class="panel-toggle" data-target="picker">−</button></div>
  <div class="body" id="pickerBody">Loading...</div>
</div>

<div id="groupbox" class="box">
  <div class="drag"><span class="drag-title">Layer toggles</span><button class="panel-toggle" data-target="groupbox">−</button></div>
  <div class="body">
    <div class="group-row">
      <input type="checkbox" id="toggleCentroidCurrent" checked />
      <label for="toggleCentroidCurrent">Show current centroid-to-centroid routes</label>
    </div>
    <div class="group-row">
      <input type="checkbox" id="toggleCentroidOptimized" checked />
      <label for="toggleCentroidOptimized">Show optimized centroid-to-centroid routes</label>
    </div>
    <div class="group-row">
      <input type="checkbox" id="toggleStreetCurrent" checked />
      <label for="toggleStreetCurrent">Show current street-to-street routes</label>
    </div>
    <div class="group-row">
      <input type="checkbox" id="toggleStreetOptimized" checked />
      <label for="toggleStreetOptimized">Show optimized street-to-street simulated routes</label>
    </div>
    <hr>
    <div class="group-row">
      <input type="checkbox" id="toggleCentroidCenters" />
      <label for="toggleCentroidCenters">Show centroid centers</label>
    </div>
    <hr>
    <div class="group-row">
      <input type="checkbox" id="hideSilver315" />
      <label for="hideSilver315">Hide Current Street | CityLink SILVER | COLUMBIA &amp; SILVER SPRING - DC | 315</label>
    </div>
    <div class="note">These toggles work on top of the individual Leaflet layer controls.</div>
  </div>
</div>

<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js" crossorigin=""></script>
<script>
const TRACTS_URL = "__TRACTS_URL__";
const CURRENT_CENTROID_GEOJSON = __CURRENT_CENTROID_GEOJSON__;
const OPTIMIZED_CENTROID_GEOJSON = __OPTIMIZED_CENTROID_GEOJSON__;
const CURRENT_STREET_GEOJSON = __CURRENT_STREET_GEOJSON__;
const OPTIMIZED_STREET_GEOJSON = __OPTIMIZED_STREET_GEOJSON__;
const CENTROID_CENTERS_GEOJSON = __CENTROID_CENTERS_GEOJSON__;

const map = L.map('map').setView([39.29,-76.61], 11);

const basemapLight = L.tileLayer('https://{s}.basemaps.cartocdn.com/light_nolabels/{z}/{x}/{y}{r}.png', {
  attribution: '&copy; OpenStreetMap &copy; CARTO', subdomains:'abcd', maxZoom:20
}).addTo(map);
const basemapOSM = L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
  attribution: '&copy; OpenStreetMap contributors', maxZoom:19
});
const basemapVoyager = L.tileLayer('https://{s}.basemaps.cartocdn.com/rastertiles/voyager/{z}/{x}/{y}{r}.png', {
  attribution: '&copy; OpenStreetMap &copy; CARTO', subdomains:'abcd', maxZoom:20
});
const basemapDark = L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png', {
  attribution: '&copy; OpenStreetMap &copy; CARTO', subdomains:'abcd', maxZoom:20
});
const basemapPositronLabels = L.tileLayer('https://{s}.basemaps.cartocdn.com/light_only_labels/{z}/{x}/{y}{r}.png', {
  attribution: '&copy; OpenStreetMap &copy; CARTO', subdomains:'abcd', maxZoom:20
});
const basemapImagery = L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', {
  attribution: 'Tiles &copy; Esri', maxZoom:19
});
const basemapTopo = L.tileLayer('https://{s}.tile.opentopomap.org/{z}/{x}/{y}.png', {
  attribution: '&copy; OpenTopoMap contributors', maxZoom:17
});
const basemapEsriStreet = L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Street_Map/MapServer/tile/{z}/{y}/{x}', {
  attribution: 'Tiles &copy; Esri', maxZoom:19
});
const basemapNatGeo = L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/NatGeo_World_Map/MapServer/tile/{z}/{y}/{x}', {
  attribution: 'Tiles &copy; Esri', maxZoom:16
});

const routeRegistry = new Map();
let centroidCentersLayer = null;
const legendBody = document.getElementById('legendBody');
const pickerBody = document.getElementById('pickerBody');

function makeDraggable(el){
  const handle = el.querySelector('.drag') || el;
  let dragging = false, sx=0, sy=0, sl=0, st=0;
  handle.addEventListener('mousedown', (e)=>{
    if (e.target.classList.contains('panel-toggle')) return;
    const r = el.getBoundingClientRect();
    dragging = true; sx = e.clientX; sy = e.clientY; sl = r.left; st = r.top;
    el.style.left = `${r.left}px`; el.style.top = `${r.top}px`; el.style.right='auto'; el.style.bottom='auto';
    e.preventDefault();
  });
  window.addEventListener('mousemove', (e)=>{
    if(!dragging) return;
    el.style.left = `${Math.max(0, sl + e.clientX - sx)}px`;
    el.style.top = `${Math.max(0, st + e.clientY - sy)}px`;
  });
  window.addEventListener('mouseup', ()=> dragging = false);
}

function bindPanelMinimizeButtons() {
  document.querySelectorAll('.panel-toggle').forEach(btn => {
    btn.addEventListener('click', () => {
      const targetId = btn.getAttribute('data-target');
      const panel = document.getElementById(targetId);
      if (!panel) return;
      panel.classList.toggle('minimized');
      btn.textContent = panel.classList.contains('minimized') ? '+' : '−';
    });
  });
}

async function fetchJson(url){
  const r = await fetch(url);
  if(!r.ok) throw new Error(`Failed to fetch ${url}`);
  return await r.json();
}

function makePopupHtml(name, p) {
  let popup = `<b>${name}</b>`;
  if (p.route_name_raw || p.route_number_raw) {
    popup += `<br>Route name: ${p.route_name_raw || ''}`;
    popup += `<br>Route number: ${p.route_number_raw || ''}`;
  }
  if (p.origin_geoid || p.ending_geoid) {
    popup += `<br>Origin tract: ${p.origin_geoid || ''}<br>Destination tract: ${p.ending_geoid || ''}`;
  }
  if (p.num_tracts) {
    popup += `<br>Tracts in route: ${p.num_tracts}`;
  }
  return popup;
}

function addRouteLayerGroup(collection, groupName, dashArray, weight, opacity){
  const overlay = {};
  (collection.features || []).forEach(feature=>{
    const props = feature.properties || {};
    const name = props.route_header || `Unknown`;
    const color = props.display_color || '#555555';
    const layer = L.geoJSON(feature, {
      style: {color: color, weight: weight, opacity: opacity, dashArray: dashArray || null},
      onEachFeature: (feat, lyr) => {
        lyr.bindPopup(makePopupHtml(name, feat.properties || {}));
      }
    });
    overlay[name] = layer;
    routeRegistry.set(name, {layer, color, dashArray, group: groupName});
  });
  return overlay;
}

function buildCentroidCentersLayer(collection){
  centroidCentersLayer = L.geoJSON(collection, {
    pointToLayer: (feature, latlng) => L.circleMarker(latlng, {
      radius: 4,
      color: '#111111',
      weight: 1,
      opacity: 1,
      fillColor: '#ffffff',
      fillOpacity: 1
    }),
    onEachFeature: (feature, layer) => {
      const p = feature.properties || {};
      const tractId = p.tract_id || 'Unknown tract';
      layer.bindPopup(`<b>Centroid center</b><br>Tract ID: ${tractId}`);
      layer.bindTooltip(`Centroid: ${tractId}`, {sticky: true});
    }
  });
  return centroidCentersLayer;
}

function setGroupVisibility(groupName, visible){
  routeRegistry.forEach((meta) => {
    if (meta.group !== groupName) return;
    if (visible) {
      if (!map.hasLayer(meta.layer)) meta.layer.addTo(map);
    } else {
      if (map.hasLayer(meta.layer)) map.removeLayer(meta.layer);
    }
  });
}

function setCentroidCentersVisibility(visible){
  if (!centroidCentersLayer) return;
  if (visible) {
    if (!map.hasLayer(centroidCentersLayer)) centroidCentersLayer.addTo(map);
  } else {
    if (map.hasLayer(centroidCentersLayer)) map.removeLayer(centroidCentersLayer);
  }
}

function setSpecificStreetRouteVisibility(predicate, visible){
  routeRegistry.forEach((meta, name) => {
    if (meta.group !== 'street_current') return;
    if (!predicate(name, meta)) return;
    if (visible) {
      if (!map.hasLayer(meta.layer)) meta.layer.addTo(map);
    } else {
      if (map.hasLayer(meta.layer)) map.removeLayer(meta.layer);
    }
  });
}

function isSilver315(name) {
  const upper = String(name || '').toUpperCase();
  return upper.includes('CURRENT STREET | CITYLINK SILVER') &&
         upper.includes('COLUMBIA & SILVER SPRING - DC') &&
         upper.includes('315');
}

function buildLegend() {
  let html = `<div><span class="swatch" style="background:#E6E6E6"></span>Tract polygons</div>`;
  html += `<div><span class="route-line" style="border-top-color:#000"></span>Current centroid-to-centroid</div>`;
  html += `<div><span class="route-line" style="border-top-color:#000;border-top-style:dashed"></span>Optimized centroid-to-centroid</div>`;
  html += `<div><span class="route-line" style="border-top-color:#000"></span>Current street-to-street</div>`;
  html += `<div><span class="route-line" style="border-top-color:#000;border-top-style:dashed"></span>Optimized street-to-street simulated</div>`;
  html += `<div><span class="centroid-dot"></span>Centroid center</div><br>`;
  [...routeRegistry.entries()].sort((a,b)=>a[0].localeCompare(b[0])).forEach(([name, meta])=>{
    const style = meta.dashArray ? 'border-top-style:dashed;' : '';
    html += `<div><span class="route-line" style="border-top-color:${meta.color};${style}"></span>${name}</div>`;
  });
  legendBody.innerHTML = html;
}

function buildPicker(){
  let html = '';
  [...routeRegistry.entries()].sort((a,b)=>a[0].localeCompare(b[0])).forEach(([name, meta], idx)=>{
    html += `<div class="picker-row"><label for="p_${idx}" title="${name}">${name}</label><input type="color" id="p_${idx}" value="${meta.color}"></div>`;
  });
  pickerBody.innerHTML = html;
  [...routeRegistry.entries()].sort((a,b)=>a[0].localeCompare(b[0])).forEach(([name, meta], idx)=>{
    document.getElementById(`p_${idx}`).addEventListener('input', (e)=>{
      meta.color = e.target.value;
      meta.layer.eachLayer(l=>{
        if(l.setStyle){ l.setStyle({color: meta.color, dashArray: meta.dashArray || null}); }
      });
      buildLegend();
    });
  });
}

async function init(){
  const tracts = await fetchJson(TRACTS_URL);
  const tractLayer = L.geoJSON(tracts, {
    style: {color:'#808080', weight:1, opacity:0.85, fillColor:'#E6E6E6', fillOpacity:0.55},
    onEachFeature: (feature, layer) => {
      const props = feature.properties || {};
      const tractId = props.GEOID || props.GEOID20 || props.GEOID10 || props.TRACTCE || 'Unknown tract';
      layer.bindTooltip(`Tract ID: ${tractId}`, {sticky:true});
    }
  }).addTo(map);

  const baseMaps = {
    'Light': basemapLight,
    'OpenStreetMap': basemapOSM,
    'Voyager': basemapVoyager,
    'Dark': basemapDark,
    'Positron Labels': basemapPositronLabels,
    'Imagery': basemapImagery,
    'Topo': basemapTopo,
    'Esri Street': basemapEsriStreet,
    'NatGeo': basemapNatGeo
  };

  const overlays = {'Tracts': tractLayer};
  Object.assign(overlays, addRouteLayerGroup(CURRENT_CENTROID_GEOJSON, 'centroid_current', null, 3.5, 0.95));
  Object.assign(overlays, addRouteLayerGroup(OPTIMIZED_CENTROID_GEOJSON, 'centroid_optimized', '8 6', 3.5, 0.95));
  Object.assign(overlays, addRouteLayerGroup(CURRENT_STREET_GEOJSON, 'street_current', null, 4.5, 0.96));
  Object.assign(overlays, addRouteLayerGroup(OPTIMIZED_STREET_GEOJSON, 'street_optimized', '8 6', 4.5, 0.96));
  const centroidLayer = buildCentroidCentersLayer(CENTROID_CENTERS_GEOJSON);
  overlays['Centroid centers'] = centroidLayer;

  L.control.layers(baseMaps, overlays, {collapsed:true}).addTo(map);

  const bounds = tractLayer.getBounds();
  if(bounds.isValid()) map.fitBounds(bounds.pad(0.03));

  buildLegend();
  buildPicker();

  setGroupVisibility('centroid_current', document.getElementById('toggleCentroidCurrent').checked);
  setGroupVisibility('centroid_optimized', document.getElementById('toggleCentroidOptimized').checked);
  setGroupVisibility('street_current', document.getElementById('toggleStreetCurrent').checked);
  setGroupVisibility('street_optimized', document.getElementById('toggleStreetOptimized').checked);
  setCentroidCentersVisibility(document.getElementById('toggleCentroidCenters').checked);
  if (document.getElementById('hideSilver315').checked) {
    setSpecificStreetRouteVisibility(isSilver315, false);
  }

  document.getElementById('toggleCentroidCurrent').addEventListener('change', (e) => {
    setGroupVisibility('centroid_current', e.target.checked);
  });
  document.getElementById('toggleCentroidOptimized').addEventListener('change', (e) => {
    setGroupVisibility('centroid_optimized', e.target.checked);
  });
  document.getElementById('toggleStreetCurrent').addEventListener('change', (e) => {
    setGroupVisibility('street_current', e.target.checked);
    if (document.getElementById('hideSilver315').checked) {
      setSpecificStreetRouteVisibility(isSilver315, false);
    }
  });
  document.getElementById('toggleStreetOptimized').addEventListener('change', (e) => {
    setGroupVisibility('street_optimized', e.target.checked);
  });
  document.getElementById('toggleCentroidCenters').addEventListener('change', (e) => {
    setCentroidCentersVisibility(e.target.checked);
  });
  document.getElementById('hideSilver315').addEventListener('change', (e) => {
    setSpecificStreetRouteVisibility(isSilver315, !e.target.checked);
  });
}

makeDraggable(document.getElementById('legend'));
makeDraggable(document.getElementById('picker'));
makeDraggable(document.getElementById('groupbox'));
makeDraggable(document.getElementById('title'));
bindPanelMinimizeButtons();
init().catch(err => console.error(err));
</script>
</body>
</html>
"""

def write_html(output_path: Path, current_geojson: dict[str, Any], optimized_geojson: dict[str, Any]) -> None:
    params = {
        "where": "1=1",
        "outFields": "*",
        "returnGeometry": "true",
        "f": "geojson",
        "outSR": "4326",
    }
    tracts_url = f"{TRACTS_URL_BASE}/query?{urllib.parse.urlencode(params)}"
    html = HTML_TEMPLATE.replace("__TRACTS_URL__", tracts_url)
    html = html.replace("__CURRENT_GEOJSON__", json.dumps(current_geojson))
    html = html.replace("__OPTIMIZED_GEOJSON__", json.dumps(optimized_geojson))
    output_path.write_text(html, encoding="utf-8")


def write_street_html(
    output_path: Path,
    current_centroid_geojson: dict[str, Any],
    optimized_centroid_geojson: dict[str, Any],
    current_street_geojson: dict[str, Any],
    optimized_simulated_geojson: dict[str, Any],
    centroid_centers_geojson: dict[str, Any],
) -> None:
    params = {
        "where": "1=1",
        "outFields": "*",
        "returnGeometry": "true",
        "f": "geojson",
        "outSR": "4326",
    }
    tracts_url = f"{TRACTS_URL_BASE}/query?{urllib.parse.urlencode(params)}"
    html = STREET_HTML_TEMPLATE.replace("__TRACTS_URL__", tracts_url)
    html = html.replace("__CURRENT_CENTROID_GEOJSON__", json.dumps(current_centroid_geojson))
    html = html.replace("__OPTIMIZED_CENTROID_GEOJSON__", json.dumps(optimized_centroid_geojson))
    html = html.replace("__CURRENT_STREET_GEOJSON__", json.dumps(current_street_geojson))
    html = html.replace("__OPTIMIZED_STREET_GEOJSON__", json.dumps(optimized_simulated_geojson))
    html = html.replace("__CENTROID_CENTERS_GEOJSON__", json.dumps(centroid_centers_geojson))
    output_path.write_text(html, encoding="utf-8")


def write_workbook(current_matrix_df: pd.DataFrame, optimized_matrix_df: pd.DataFrame, output_xlsx: Path) -> None:
    if OPENPYXL_AVAILABLE:
        engine = "openpyxl"
    elif XLSXWRITER_AVAILABLE:
        engine = "xlsxwriter"
    else:
        return

    with pd.ExcelWriter(output_xlsx, engine=engine) as writer:
        current_matrix_df.to_excel(writer, sheet_name="current_route_matrix", index=False)
        optimized_matrix_df.to_excel(writer, sheet_name="optimized_route_matrix", index=False)

    if not OPENPYXL_AVAILABLE:
        return

    wb = load_workbook(output_xlsx)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "B2"
        ws.sheet_view.showGridLines = False
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.fill = PatternFill(fill_type="solid", fgColor="D9EAF7" if col > 1 else "D9D9D9")
            ws.column_dimensions[get_column_letter(col)].width = 14 if col > 1 else 42
        for row in range(2, ws.max_row + 1):
            route_label = str(ws.cell(row=row, column=1).value or "")
            fill = route_color(route_label).replace("#", "")
            ws.cell(row=row, column=1).fill = PatternFill(fill_type="solid", fgColor=fill)
            ws.cell(row=row, column=1).font = Font(bold=True)
    wb.save(output_xlsx)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ask for bussing_times distance matrix and current/optimized route matrices, then create 3 centroid-to-centroid PNGs and an HTML map.")
    parser.add_argument("--interactive", action="store_true", help="Ask for paths in the terminal.")
    parser.add_argument("--distance-matrix", type=Path, default=None, help="bussing_times CSV/XLSX (minutes distance matrix).")
    parser.add_argument("--current-route", type=Path, default=None, help="current route matrix (tractroute compact).")
    parser.add_argument("--optimized-route", type=Path, default=None, help="optimized route matrix. Optional; if omitted, optimized routes are computed from the distance matrix.")
    parser.add_argument("--output-dir", type=Path, default=None, help="Output folder. Defaults next to the distance matrix.")
    return parser.parse_args()

def main() -> None:
    args = parse_args()

    default_distance = Path(r"C:\Users\benac\Downloads\bussing_times.csv")
    default_current = Path(r"C:\Users\benac\Downloads\tractbyroutefinalcompact.xlsx")

    if args.interactive or args.distance_matrix is None:
        distance_matrix_path = ask_path("Enter the path to bussing_times (this is the distance matrix)", default_distance)
    else:
        distance_matrix_path = args.distance_matrix

    if args.interactive or args.current_route is None:
        current_route_path = ask_path("Enter the path to the current route matrix (tractroute compact)", default_current)
    else:
        current_route_path = args.current_route

    if args.interactive or args.optimized_route is None:
        optimized_route_path = ask_path(
            "Enter the path to the optimized route matrix (press Enter to auto-build from the distance matrix)",
            None,
            allow_blank=True,
        )
    else:
        optimized_route_path = args.optimized_route

    default_output_dir = distance_matrix_path.parent / f"{distance_matrix_path.stem}_centroid_route_outputs"
    if args.interactive:
        output_dir = ask_output_dir("Enter the output folder", default_output_dir)
    else:
        output_dir = args.output_dir if args.output_dir is not None else default_output_dir
        ensure_directory(output_dir)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    current_route_df, current_tract_cols, current_sheet = load_route_matrix(current_route_path)
    distance_df = load_distance_matrix(distance_matrix_path)

    current_bundle_df = build_route_bundle(current_route_df, current_tract_cols, "Current")

    if optimized_route_path is not None:
        optimized_route_df, optimized_tract_cols, optimized_sheet = load_route_matrix(optimized_route_path)
        tract_cols = optimized_tract_cols
        optimized_bundle_df = build_route_bundle(optimized_route_df, optimized_tract_cols, "Optimized")
    else:
        tract_cols = current_tract_cols
        optimized_sheet = ""
        optimized_bundle_df = compute_optimized_from_distance(current_bundle_df, current_tract_cols, distance_df)

    current_bundle_df, optimized_bundle_df = align_bundle_route_sets(current_bundle_df, optimized_bundle_df)

    centroid_lookup, tract_polys = load_tract_geometries()

    current_bundle_df = filter_bundle_to_baltimore(current_bundle_df, tract_cols, centroid_lookup)
    optimized_bundle_df = filter_bundle_to_baltimore(optimized_bundle_df, tract_cols, centroid_lookup)

    current_png = output_dir / "current_routes_centroid_to_centroid.png"
    optimized_png = output_dir / "optimized_routes_centroid_to_centroid.png"
    combo_png = output_dir / "current_and_optimized_routes_combo.png"

    current_png_more_opaque = output_dir / "current_routes_centroid_to_centroid_more_opaque.png"
    optimized_png_more_opaque = output_dir / "optimized_routes_centroid_to_centroid_more_opaque.png"
    combo_png_more_opaque = output_dir / "current_and_optimized_routes_combo_more_opaque.png"

    centroid_combo_with_legend_png = output_dir / "current_and_optimized_routes_combo_centroid_to_centroid_with_legend.png"
    street_combo_with_legend_png = output_dir / "current_and_optimized_routes_combo_street_to_street_with_legend.png"

    current_live_street_png = output_dir / "current_routes_live_mta_street_level.png"
    current_live_street_png_no_silver_dc = output_dir / "current_routes_live_mta_street_level_no_silver_dc.png"
    optimized_simulated_png = output_dir / "optimized_routes_simulated_following_centroid_path.png"

    html_path = output_dir / "Baltimore CityLink Centroid Routes.html"
    street_html_path = output_dir / "Baltimore CityLink Street-Level Current vs Simulated Optimized Routes.html"

    current_live_street_geojson_path = output_dir / "current_routes_live_mta_street_level.geojson"
    optimized_simulated_geojson_path = output_dir / "optimized_routes_simulated_following_centroid_path.geojson"
    centroid_centers_geojson_path = output_dir / "centroid_centers.geojson"

    optimized_more_opaque_route_audit_path = output_dir / "optimized_routes_centroid_to_centroid_more_opaque_route_audit.csv"
    route_count_audit_path = output_dir / "current_vs_optimized_route_count_audit.csv"

    workbook_path = output_dir / "citylink_current_and_optimized_route_matrices.xlsx"
    current_geojson_path = output_dir / "current_routes_centroid.geojson"
    optimized_geojson_path = output_dir / "optimized_routes_centroid.geojson"
    metadata_path = output_dir / "run_metadata.json"

    plot_bundle_png(
        current_bundle_df,
        tract_cols,
        centroid_lookup,
        tract_polys,
        current_png,
        "Current CityLink routes\ntract centroid-to-centroid",
        optimized_style=False,
        route_alpha=0.95,
    )

    plot_bundle_png(
        optimized_bundle_df,
        tract_cols,
        centroid_lookup,
        tract_polys,
        optimized_png,
        "Optimized CityLink routes\ntract centroid-to-centroid",
        optimized_style=True,
        route_alpha=0.95,
    )

    plot_combo_png(
        current_bundle_df,
        optimized_bundle_df,
        tract_cols,
        centroid_lookup,
        tract_polys,
        combo_png,
        route_alpha=0.92,
    )

    plot_bundle_png(
        current_bundle_df,
        tract_cols,
        centroid_lookup,
        tract_polys,
        current_png_more_opaque,
        "Current CityLink routes (more opaque)\ntract centroid-to-centroid with centroid dots",
        optimized_style=False,
        route_alpha=0.72,
        show_centroid_dots=True,
        show_route_legend=True,
    )

    optimized_route_audit_df, optimized_missing_routes = audit_plottable_routes(
        optimized_bundle_df, tract_cols, centroid_lookup
    )
    optimized_route_audit_df.to_csv(optimized_more_opaque_route_audit_path, index=False)

    if optimized_missing_routes:
        raise ValueError(
            "The following optimized routes would be missing from "
            f"'optimized_routes_centroid_to_centroid_more_opaque.png': {optimized_missing_routes}. "
            f"See audit file: {optimized_more_opaque_route_audit_path}"
        )

    plot_bundle_png(
        optimized_bundle_df,
        tract_cols,
        centroid_lookup,
        tract_polys,
        optimized_png_more_opaque,
        "Optimized CityLink routes (more opaque)\ntract centroid-to-centroid with centroid dots",
        optimized_style=True,
        route_alpha=0.72,
        show_centroid_dots=True,
        show_route_legend=True,
    )

    plot_combo_png(
        current_bundle_df,
        optimized_bundle_df,
        tract_cols,
        centroid_lookup,
        tract_polys,
        combo_png_more_opaque,
        route_alpha=0.72,
        show_centroid_dots=True,
        show_route_legend=True,
    )

    plot_combo_png(
        current_bundle_df,
        optimized_bundle_df,
        tract_cols,
        centroid_lookup,
        tract_polys,
        centroid_combo_with_legend_png,
        route_alpha=0.72,
        show_centroid_dots=True,
        show_route_legend=True,
    )

    plot_live_mta_current_routes_png(
        current_bundle_df,
        tract_polys,
        current_live_street_png,
        "Current CityLink routes pulled from live MTA street-level geometry",
    )

    plot_live_mta_current_routes_png(
        current_bundle_df,
        tract_polys,
        current_live_street_png_no_silver_dc,
        "Current CityLink routes pulled from live MTA street-level geometry (without Silver DC 315)",
        exclude_silver_315=True,
    )

    plot_simulated_optimized_routes_png(
        optimized_bundle_df,
        tract_cols,
        centroid_lookup,
        tract_polys,
        optimized_simulated_png,
        "Road-following optimized CityLink routes snapped to streets from the optimized centroid-to-centroid path",
    )

    plot_street_combo_png(
        current_bundle_df,
        optimized_bundle_df,
        tract_cols,
        centroid_lookup,
        tract_polys,
        street_combo_with_legend_png,
        "Baltimore street-to-street routes",
        route_alpha=0.92,
        show_route_legend=True,
    )

    current_geojson = features_from_bundle(current_bundle_df, tract_cols, centroid_lookup, optimized_style=False)
    optimized_geojson = features_from_bundle(optimized_bundle_df, tract_cols, centroid_lookup, optimized_style=True)

    current_geojson_path.write_text(json.dumps(current_geojson, indent=2), encoding="utf-8")
    optimized_geojson_path.write_text(json.dumps(optimized_geojson, indent=2), encoding="utf-8")
    write_html(html_path, current_geojson, optimized_geojson)

    current_live_street_geojson = build_live_mta_current_geojson(current_bundle_df)
    optimized_simulated_geojson = build_simulated_optimized_geojson(optimized_bundle_df, tract_cols, centroid_lookup)

    route_count_audit_df, route_count_errors = validate_current_optimized_route_counts(
        current_bundle_df,
        optimized_bundle_df,
        current_live_street_geojson,
        optimized_simulated_geojson,
    )
    route_count_audit_df.to_csv(route_count_audit_path, index=False)

    if route_count_errors:
        raise ValueError(
            "Current vs optimized route counts/sets do not match. "
            + " | ".join(route_count_errors)
            + f" See audit file: {route_count_audit_path}"
        )

    centroid_centers_geojson = build_centroid_centers_geojson(
        current_bundle_df,
        optimized_bundle_df,
        tract_cols,
        centroid_lookup,
    )

    current_live_street_geojson_path.write_text(json.dumps(current_live_street_geojson, indent=2), encoding="utf-8")
    optimized_simulated_geojson_path.write_text(json.dumps(optimized_simulated_geojson, indent=2), encoding="utf-8")
    centroid_centers_geojson_path.write_text(json.dumps(centroid_centers_geojson, indent=2), encoding="utf-8")

    write_street_html(
        street_html_path,
        current_geojson,
        optimized_geojson,
        current_live_street_geojson,
        optimized_simulated_geojson,
        centroid_centers_geojson,
    )

    current_matrix_df = matrix_for_excel(current_bundle_df, tract_cols)
    optimized_matrix_df = matrix_for_excel(optimized_bundle_df, tract_cols)
    write_workbook(current_matrix_df, optimized_matrix_df, workbook_path)

    metadata = {
        "distance_matrix": str(distance_matrix_path),
        "current_route_input": str(current_route_path),
        "current_sheet_used": current_sheet,
        "optimized_route_input": str(optimized_route_path) if optimized_route_path else None,
        "optimized_sheet_used": optimized_sheet,
        "optimized_source": (
            "supplied optimized route matrix"
            if optimized_route_path
            else "computed from distance matrix using shortest path between current route origin and destination"
        ),
        "assumptions": {
            "bussing_times_meaning": "distance matrix in minutes",
            "zero_meaning": "0 = no direct connection",
            "current_route_file": "tractroute compact current route matrix",
            "html_optimized_routes": "toggleable, color-changeable optimized routes passing through tract centroid centers",
            "baltimore_filter": "all current and optimized route bundles are filtered to keep only routes whose centroid path is at least somewhat in Baltimore",
        },
        "outputs": {
            "current_png": str(current_png),
            "optimized_png": str(optimized_png),
            "combo_png": str(combo_png),
            "current_png_more_opaque": str(current_png_more_opaque),
            "optimized_png_more_opaque": str(optimized_png_more_opaque),
            "combo_png_more_opaque": str(combo_png_more_opaque),
            "centroid_combo_with_legend_png": str(centroid_combo_with_legend_png),
            "street_combo_with_legend_png": str(street_combo_with_legend_png),
            "current_live_street_png": str(current_live_street_png),
            "current_live_street_png_no_silver_dc": str(current_live_street_png_no_silver_dc),
            "optimized_simulated_png": str(optimized_simulated_png),
            "html": str(html_path),
            "street_html": str(street_html_path),
            "current_live_street_geojson": str(current_live_street_geojson_path),
            "optimized_simulated_geojson": str(optimized_simulated_geojson_path),
            "centroid_centers_geojson": str(centroid_centers_geojson_path),
            "optimized_more_opaque_route_audit_csv": str(optimized_more_opaque_route_audit_path),
            "route_count_audit_csv": str(route_count_audit_path),
            "workbook": str(workbook_path),
            "current_geojson": str(current_geojson_path),
            "optimized_geojson": str(optimized_geojson_path),
        },
        "counts": {
            "current_routes": int(len(current_bundle_df)),
            "optimized_routes": int(len(optimized_bundle_df)),
        },
    }

    metadata_path.write_text(json.dumps(metadata, indent=2), encoding="utf-8")

    print(f"Using distance matrix: {distance_matrix_path}")
    print(f"Using current route matrix: {current_route_path}")
    if optimized_route_path:
        print(f"Using optimized route matrix: {optimized_route_path}")
    else:
        print("No optimized route matrix provided; optimized routes were computed from the distance matrix.")

    print(f"Output folder: {output_dir}")
    print(f"Saved current-only centroid PNG: {current_png}")
    print(f"Saved optimized-only centroid PNG: {optimized_png}")
    print(f"Saved combo centroid PNG: {combo_png}")
    print(f"Saved current-only more-opaque centroid PNG: {current_png_more_opaque}")
    print(f"Saved optimized-only more-opaque centroid PNG: {optimized_png_more_opaque}")
    print(f"Saved combo more-opaque centroid PNG: {combo_png_more_opaque}")
    print(f"Saved centroid-to-centroid combo PNG with legend: {centroid_combo_with_legend_png}")
    print(f"Saved street-to-street combo PNG with legend: {street_combo_with_legend_png}")
    print(f"Saved current live MTA street-level PNG: {current_live_street_png}")
    print(f"Saved current live MTA street-level PNG without Silver DC 315: {current_live_street_png_no_silver_dc}")
    print(f"Saved optimized more-opaque route audit CSV: {optimized_more_opaque_route_audit_path}")
    print(f"Saved current vs optimized route count audit CSV: {route_count_audit_path}")
    print(f"Saved simulated optimized route PNG: {optimized_simulated_png}")
    print(f"Saved centroid HTML map: {html_path}")
    print(f"Saved street+centroid HTML map with current/optimized centroid and street routes plus centroid-centers toggle: {street_html_path}")
    print(f"Saved workbook: {workbook_path}")
    print(f"Saved metadata: {metadata_path}")

if __name__ == "__main__":
    main()